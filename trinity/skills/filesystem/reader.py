"""
Trinity Skill — File System Reader.
Reads files and directories.
"""

import os
import structlog
from pathlib import Path
from datetime import datetime
from trinity.skills.base import BaseSkill, SkillResult

logger = structlog.get_logger(__name__)

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB


class FileSystemReader(BaseSkill):
    """Read files and list directories."""

    async def execute(self, entities: dict, context: dict | None = None) -> SkillResult:
        """Execute read operation based on entities."""
        raw_text = entities.get("raw_text", "")

        # Determine if this is a list_dir or read_file
        if any(kw in raw_text.lower() for kw in ["what's in", "what is in", "list", "show me", "contents of"]):
            return await self.list_dir(entities)
        return await self.read_file(entities)

    async def read_file(self, entities: dict) -> SkillResult:
        """Read the contents of a file."""
        path_str = entities.get("path", entities.get("raw_text", ""))
        path = Path(self._resolve_path(path_str))

        # Find the file if path is ambiguous
        if not path.exists():
            found = await self._find_file(path_str)
            if found:
                path = Path(found)
            else:
                return self._error(f"I couldn't find a file matching '{path_str}'.")

        if path.is_dir():
            return await self.list_dir(entities)

        if not path.is_file():
            return self._error(f"'{path.name}' is not a file.")

        # Check file size
        size = path.stat().st_size
        if size > MAX_FILE_SIZE:
            return self._error(
                f"This file is {size / 1024 / 1024:.1f}MB — too large to read entirely. "
                "I can show you the first few lines or search within it."
            )

        # Read based on file type
        try:
            content = await self._read_by_type(path)
            return self._success(f"Here's the content of {path.name}:\n\n{content}",
                                data={"path": str(path), "content": content})
        except PermissionError:
            return self._error(f"I don't have permission to read '{path.name}'.")
        except Exception as e:
            return self._error(f"Error reading '{path.name}': {str(e)}")

    async def list_dir(self, entities: dict) -> SkillResult:
        """List contents of a directory."""
        path_str = entities.get("path", entities.get("raw_text", ""))

        # Parse common folder names
        if not path_str or any(kw in path_str.lower() for kw in ["folder", "directory", "dir"]):
            for alias in ["desktop", "documents", "downloads", "pictures", "music", "videos"]:
                if alias in path_str.lower():
                    path_str = alias
                    break

        path = Path(self._resolve_path(path_str))

        if not path.exists():
            return self._error(f"The folder '{path_str}' doesn't exist. Should I create it?")

        if not path.is_dir():
            return await self.read_file(entities)

        try:
            items = list(path.iterdir())
            result_lines = []

            for item in sorted(items, key=lambda x: (not x.is_dir(), x.name.lower())):
                icon = "📁" if item.is_dir() else self._file_icon(item)
                size = ""
                if item.is_file():
                    try:
                        s = item.stat().st_size
                        size = f" ({self._format_size(s)})"
                    except Exception:
                        pass
                modified = ""
                try:
                    mtime = datetime.fromtimestamp(item.stat().st_mtime)
                    modified = f" — {mtime.strftime('%b %d, %I:%M %p')}"
                except Exception:
                    pass
                result_lines.append(f"{icon} {item.name}{size}{modified}")

            total = len(items)
            dirs = sum(1 for i in items if i.is_dir())
            files = total - dirs

            header = f"Contents of {path.name} ({dirs} folders, {files} files):\n\n"
            content = header + "\n".join(result_lines)

            return self._success(content, data={"path": str(path), "items": [i.name for i in items]})

        except PermissionError:
            return self._error(f"I don't have permission to access '{path.name}'.")
        except Exception as e:
            return self._error(f"Error listing '{path.name}': {str(e)}")

    async def _read_by_type(self, path: Path) -> str:
        """Read file content based on its type."""
        suffix = path.suffix.lower()

        if suffix in (".txt", ".md", ".csv", ".json", ".yaml", ".yml", ".log", ".ini", ".cfg", ".env"):
            return path.read_text(encoding="utf-8", errors="replace")

        if suffix in (".py", ".js", ".ts", ".html", ".css", ".xml", ".sh", ".bat", ".ps1"):
            return path.read_text(encoding="utf-8", errors="replace")

        if suffix == ".pdf":
            return await self._read_pdf(path)

        if suffix == ".docx":
            return await self._read_docx(path)

        if suffix == ".xlsx":
            return await self._read_xlsx(path)

        # Try as text
        try:
            return path.read_text(encoding="utf-8", errors="replace")
        except Exception:
            return f"[Binary file: {self._format_size(path.stat().st_size)}]"

    async def _read_pdf(self, path: Path) -> str:
        """Read text from PDF."""
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(str(path))
            text = ""
            for page in doc:
                text += page.get_text()
            doc.close()
            return text[:50000]  # Cap at 50k chars
        except ImportError:
            return "[PDF reader not available — install PyMuPDF]"
        except Exception as e:
            return f"[Error reading PDF: {str(e)}]"

    async def _read_docx(self, path: Path) -> str:
        """Read text from Word document."""
        try:
            from docx import Document
            doc = Document(str(path))
            return "\n".join([p.text for p in doc.paragraphs])
        except ImportError:
            return "[DOCX reader not available — install python-docx]"
        except Exception as e:
            return f"[Error reading DOCX: {str(e)}]"

    async def _read_xlsx(self, path: Path) -> str:
        """Read text from Excel spreadsheet."""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True)
            text = ""
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text += f"--- Sheet: {sheet_name} ---\n"
                for row in ws.iter_rows(values_only=True):
                    text += " | ".join([str(cell) if cell else "" for cell in row]) + "\n"
            wb.close()
            return text[:50000]
        except ImportError:
            return "[XLSX reader not available — install openpyxl]"
        except Exception as e:
            return f"[Error reading XLSX: {str(e)}]"

    async def _find_file(self, name: str) -> str | None:
        """Try to find a file by name in common directories."""
        search_dirs = [
            Path.home() / "Desktop",
            Path.home() / "Documents",
            Path.home() / "Downloads",
            Path.home(),
        ]
        for search_dir in search_dirs:
            if not search_dir.exists():
                continue
            try:
                for item in search_dir.iterdir():
                    if name.lower() in item.name.lower():
                        return str(item)
            except PermissionError:
                continue
        return None

    @staticmethod
    def _file_icon(path: Path) -> str:
        """Get an appropriate icon for a file type."""
        icons = {
            ".pdf": "📄", ".docx": "📝", ".xlsx": "📊", ".pptx": "📊",
            ".txt": "📄", ".md": "📄", ".csv": "📊", ".json": "📄",
            ".py": "🐍", ".js": "⚡", ".html": "🌐", ".css": "🎨",
            ".jpg": "🖼️", ".jpeg": "🖼️", ".png": "🖼️", ".gif": "🖼️",
            ".mp3": "🎵", ".wav": "🎵", ".mp4": "🎬", ".avi": "🎬",
            ".zip": "📦", ".rar": "📦", ".exe": "⚙️",
        }
        return icons.get(path.suffix.lower(), "📄")

    @staticmethod
    def _format_size(size: int) -> str:
        """Format file size in human-readable format."""
        for unit in ["B", "KB", "MB", "GB"]:
            if size < 1024:
                return f"{size:.1f} {unit}"
            size /= 1024
        return f"{size:.1f} TB"
